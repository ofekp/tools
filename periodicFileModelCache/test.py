import subprocess
import sys
import re
import urllib2
import threading
import time
from pexpect import pxssh
import base64
import re
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl import load_workbook
from multiprocessing import Pool, Process
from subprocess import PIPE, Popen
from threading import Thread
from Queue import Queue, Empty

username = "ofek"
host = "visitor4.nads.bf1.yahoo.com"
command = "tail -f /home/y/logs/vespa/vespa.log"
hdfs_host = "phazon-gw.{color}.ygrid.yahoo.com"
# hdfs_command = "HADOOP_CLIENT_OPTS='-Xmx1024M' hadoop fs -text hdfs://phazon" + color + "-nn1." + color + ".ygrid.yahoo.com/projects/cb_native_ads/gemsci/latest/history/*201709260[78]*lfmModel112* | grep 4250754"
hdfs_mkdir_command = "hadoop fs -mkdir -p /tmp/" + username + "/test_result/"
hdfs_clean_command = "hadoop fs -rm -r /tmp/" + username + "/test_result/"
#model_number = "119"
model_number = "123"
hdfs_print_command = "hadoop fs -text /tmp/" + username + "/test_result/*"
hdfs_kinit_command = 'kinit ' + username + '@Y.CORP.YAHOO.COM'

pattern = re.compile(r'(\d*\.\d*)\s.*, .*beta=(.*), alpha=(.*)\}\].*')
timeout_sec = 20
time_between_checks = 30
#bucket_names = ["copyProdMain", "sonTurnOnPeriodicFileModelCache"]
bucket_names = ["son30MinTrainingIntervalFileCache", "son30MinTrainingInterval"]
search_string = "[SearchPctrModifiers{feature=section, label=4250754"

model_timestamp_pattern = re.compile(r'.*/modelFile_main_(\d*)_.*\s(\{.*\})')
# model_current_pattern = re.compile('.*\s(\{.*\})')
model_line_pattern = re.compile(r'.*\/modelFile_main_(\d*)_.*\s\{.*4250754\",\s*"alpha":\s*([0-9-.]*),\s*"beta":\s*([0-9-.]*)\}')
model_ts_map = {}

now = datetime.now() - timedelta(hours=3)  # UTC time
xlsx_file_ts = "%d%02d%02d%02d%02d" % (now.year, now.month, now.day, now.hour, now.minute)
xlsx_file = "test_" + xlsx_file_ts + ".csv"


def update_son_models_history(color):
    print("Updating son models history from HDFS")
    now = datetime.now() - timedelta(hours=3)  # UTC time
    day_now = "%d%02d%02d" % (now.year, now.month, now.day)

    try:
        hdfs_ssh = pxssh.pxssh()
        hdfs_ssh.login(hdfs_host.format(color=color), "ofek")
        print("sending command [" + hdfs_kinit_command + "]")
        hdfs_ssh.sendline(hdfs_kinit_command)
        res = hdfs_ssh.expect(["Password for", pxssh.EOF, pxssh.TIMEOUT])
        if res == 0:
            with open('config', 'r') as config_file:
                base64pass = config_file.readline()
                password = base64.b64decode(base64pass)
                print("sending password")
                hdfs_ssh.sendline(password)
            hdfs_ssh.prompt(1000)

            # make a list of all the relevant latest model files
            hdfs_ls_command = "HADOOP_CLIENT_OPTS=\"-Xmx1024M\" hadoop fs -ls hdfs://phazon" + color + "-nn1." + color + ".ygrid.yahoo.com/projects/cb_native_ads/gemsci/latest/history/" + model_number + "/*" + day_now + "*lfmModel" + model_number + "* | grep \"\\-rwxrwxr\\-x\" | tail -10"
            print("sending command [" + hdfs_ls_command + "]")
            hdfs_ssh.sendline(hdfs_ls_command)
            hdfs_ssh.prompt(15000)
            print(hdfs_ssh.before)
            models_file_list = ["\"" + l.strip().split()[7] + "\"" for l in str(hdfs_ssh.before).split('\n') if l.rstrip() and "rwxrwxr" in l and "modelFile_main_" in l]
            input_files_as_params = "-input "
            input_files_as_params += " -input ".join(models_file_list)

            # add model from 'current' dir
            # input_files_as_params += " -input \"hdfs://phazon" + color + "-nn1." + color + ".ygrid.yahoo.com/projects/cb_native_ads/gemsci/latest/sonModels/lfmModel112/main/*\""

            hdfs_command = "HADOOP_CLIENT_OPTS=\"-Xmx1024M\" hadoop jar /home/gs/hadoop/current/share/hadoop/tools/lib/hadoop-streaming.jar -D mapred.reduce.tasks=0 -D stream.tmpdir=\"/tmp\" " + input_files_as_params + " -output \"/tmp/" + username + "/test_result/\" -mapper \"python -c \\\"import sys, os;print('\\n'.join([os.environ['mapreduce_map_input_file']+' '+s for s in sys.stdin.readlines() if s.find('4250754')>-1]))\\\"\""

            print("sending command [" + hdfs_mkdir_command + "]")
            hdfs_ssh.sendline(hdfs_mkdir_command)
            hdfs_ssh.prompt(3000)
            print("sending command [" + hdfs_clean_command + "]")
            hdfs_ssh.sendline(hdfs_clean_command)
            hdfs_ssh.prompt(3000)
            print("sending command [" + hdfs_command + "]")
            hdfs_ssh.sendline(hdfs_command)
            hdfs_ssh.prompt(40000)
            print("sending command [" + hdfs_print_command + "]")
            hdfs_ssh.sendline(hdfs_print_command)
            hdfs_ssh.prompt(30000)

            # for debugging use:
            # with open("hdfs_debug_output", "r") as file:
            # lines = [l.strip() for l in file.readlines() if l.rstrip()]
            print(hdfs_ssh.before)
            print(str(str(hdfs_ssh.before).split('\n')))
            lines = [l.strip() for l in str(hdfs_ssh.before).split('\n') if l.rstrip()]
            for line in lines:
                # if "hdfs://phazon" + color + "-nn1." + color + ".ygrid.yahoo.com/projects/cb_native_ads/gemsci/latest/sonModels/lfmModel112/main/" in line:
                #     m = re.match(model_current_pattern, line)
                #     if m is not None:
                #         model_ts_map[m.group(1).replace(" ", "")] = "current (" + color + ")"
                #     else:
                #         print("error while extracting ts from current model file")
                # else:
                print(line)
                m = re.match(model_line_pattern, line)
                if m is not None:
                    print(str(m.group(1)) + " (" + color + ") ==> alpha [" + m.group(2) + "] beta [" + m.group(3) + "]")
                    model_ts_map[(m.group(2), m.group(3))] = str(m.group(1)) + " (" + color + ")"
                else:
                    print("error while extracting ts from model file")
        else:
            print("Error while sending password to host [" + hdfs_host.format(color=color) + "]")
    except Exception as e:
        print(e)

    print("SoN models history updated")
    print(str(model_ts_map))

    # with open(xlsx_file, 'w') as f:
    #     f.write("\n")
    #     for i, t in enumerate(model_ts_map.items()):
    #         ws2['A' + str(i)] = str(t[0])
    #         ws2['B' + str(i)] = str(t[1])
    #     wb2.save(filename=xlsx_file)
    # initiate this method every 10 minutes

    # threading.Timer(600, update_son_models_history, [color]).start()


def make_call(bucket):
    print("executing call")
    with open("profile", "r") as profile_file:
        data = profile_file.readline()
        cont_len = len(data)
        req = urllib2.Request("http://" + host + ":4080/search/?debug=true&explicitSecId=4250754&queryProfile=curveball&bucketId=" + bucket + "&useOrganicFederations=true&useDPA=false&enableRtb=false&presentation.format=xml&timeout=10000", data, {'Content-Type': 'application/x-www-form-urlencoded', 'Content-Length': cont_len})
        f = urllib2.urlopen(req)
        response = f.read()
        f.close()
    print("Thread done")


def enqueue_log_output(ssh, queue):
    for line in iter(ssh.stdout.readline, b''):
        queue.put(line)


wb = Workbook()
ws = wb.active
wb.create_sheet('map')
with open(xlsx_file, 'w') as f:
    f.write(','.join([' ', 'Most Recent Model'] + bucket_names) + "\n")

update_son_models_history("red")
update_son_models_history("blue")

if model_ts_map == {}:
    print("could not update map")
    exit(1)

count = 0
should_update_models_history = False
while True:
    print("---")
    # find the current know newest model
    newest_model_ts = max([long(model_ts_map[s].split()[0]) for s in model_ts_map])
    count += 1
    csv_line_out = [str(count), str(newest_model_ts)] + ["" for _ in bucket_names]
    for i, bucket_name in enumerate(bucket_names):
        t = threading.Thread(target=make_call, args=(bucket_name,))

        # [creative 33139426555] [SearchPctrModifiers{feature=section, label=4250754, beta=0.04713, alpha=-3.498046}]
        ssh = subprocess.Popen(["ssh", "%s" % host, command],
                            shell=False,
                            stdout=subprocess.PIPE,
                            stderr=subprocess.PIPE)

        ts_with_timeout = time.time() + 5
        for line in iter(ssh.stdout.readline, ''):
            # sys.stdout.write(line)
            if time.time() > ts_with_timeout:
                break

        t.start()
        t.join()

        q = Queue()
        t = Thread(target=enqueue_log_output, args=(ssh, q))
        t.daemon = True  # thread dies with the program
        t.start()

        # read line without blocking
        ts_with_timeout = time.time() + timeout_sec
        isTimeout = False
        found_line = None
        while time.time() < ts_with_timeout:
            try:
                line = q.get_nowait()  # or q.get(timeout=.1)
            except Empty:
                pass
            else:
                # got line
                if search_string in line:
                    found_line = line
                    print("--> " + found_line.strip())
                    break

        # for line in iter(ssh.stdout.readline, ''):
        #     #sys.stdout.write(line)
        #     if search_string in line:
        #         found_line = line
        #         print("--> " + found_line.strip())
        #         break
        #     if time.time() > ts_with_timeout:
        #         isTimeout = True
        #         break

        if found_line is None:
            print("bucket [" + bucket_name + "] TIMEOUT")
            csv_line_out[2 + i] = 'TIMEOUT'
        else:
            matches = pattern.findall(line)
            if not any(matches):
                print("line [" + found_line + "] is not a valid line to parse")
                csv_line_out[2 + i] = 'LINE NOT VALID'
                continue
            timestamp, beta, alpha = matches[0]  # expecting exactly one match
            alpha = '{:.6f}'.format(float(alpha))
            beta = '{:.6f}'.format(float(beta))

            map_key = (alpha, beta)
            if map_key in model_ts_map:
                print("bucket [" + bucket_name + "] map key [" + str(map_key) + "] value [" + model_ts_map[map_key] + "]")
                csv_line_out[2 + i] = model_ts_map[map_key]
            else:
                # error, could not find the model
                should_update_models_history = True
                print("bucket [" + bucket_name + "] could not find model for map key [" + str(map_key) + "]")
                csv_line_out[2 + i] = "\"" + str(map_key) + "\""

        print("")
        print("")
    with open(xlsx_file, 'a') as f:
        f.write(','.join(map(str, csv_line_out)) + "\n")
    if should_update_models_history or count % 5 == 0:
        should_update_models_history = False
        update_son_models_history("red")
        update_son_models_history("blue")
    else:
        time.sleep(time_between_checks)

